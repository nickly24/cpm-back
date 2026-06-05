"""Парсинг Excel-файла импорта пользователей."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

HEADER_ALIASES = {
    "full_name": (
        "фио",
        "фио ученика",
        "ученик",
        "студент",
        "full_name",
        "имя",
        "ф.и.о",
        "ф.и.о.",
    ),
    "class": ("класс", "class"),
    "school": ("школа", "school"),
    "proctor": ("проктор", "proctor", "куратор"),
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    normalized = [_normalize_header(cell) for cell in header_row]

    for field, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = index
                break

    return mapping


def _cell_value(row: tuple, index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def parse_users_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return {"status": False, "error": f"Не удалось прочитать Excel: {exc}"}

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return {"status": False, "error": "Файл пустой"}

    column_map = _map_headers(list(header_row))
    missing = [name for name in ("full_name", "class") if name not in column_map]
    if missing:
        workbook.close()
        return {
            "status": False,
            "error": "Не найдены обязательные колонки: ФИО и Класс. "
            "Проверьте заголовки первой строки.",
        }

    parsed_rows: List[Dict[str, Any]] = []
    row_number = 1

    for row in rows_iter:
        row_number += 1
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        full_name = _cell_value(row, column_map.get("full_name"))
        class_raw = _cell_value(row, column_map.get("class"))
        school_name = _cell_value(row, column_map.get("school"))
        proctor_name = _cell_value(row, column_map.get("proctor"))

        parsed_rows.append(
            {
                "row": row_number,
                "full_name": full_name,
                "class_raw": class_raw,
                "school_name": school_name,
                "proctor_name": proctor_name,
            }
        )

    workbook.close()

    if not parsed_rows:
        return {"status": False, "error": "В файле нет строк с данными"}

    return {"status": True, "rows": parsed_rows}
