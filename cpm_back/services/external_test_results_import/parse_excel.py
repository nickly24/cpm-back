"""Парсинг Excel-выгрузки результатов внешнего теста."""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

HEADER_SEARCH_ROWS = 10

HEADER_ALIASES = {
    "source_number": ("№", "номер", "id", "номер попытки"),
    "platform_user": ("пользователь", "user", "username"),
    "ip": ("ip", "ip адрес", "ip-адрес"),
    "completed_at": ("дата завершения", "завершено", "completed_at", "completed"),
    "time_spent": ("потрачено времени", "время", "time_spent", "duration"),
    "login": ("логин", "login"),
    "full_name": ("фио", "ф.и.о", "ф.и.о.", "ученик", "студент", "full_name"),
    "correct_count": ("количество правильных ответов", "правильных ответов", "верных ответов"),
    "percent": (
        "процент правильных ответов (%)",
        "процент правильных ответов",
        "процент",
        "rate",
        "score",
    ),
}

REQUIRED_FIELDS = ("full_name", "percent")


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_value(row: Tuple[Any, ...], index: Optional[int]) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _string_value(row: Tuple[Any, ...], index: Optional[int]) -> str:
    value = _cell_value(row, index)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _map_headers(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    normalized = [_normalize_header(cell) for cell in header_row]
    mapping: Dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        for index, header in enumerate(normalized):
            if header in normalized_aliases:
                mapping[field] = index
                break
    return mapping


def _find_header(rows: List[Tuple[Any, ...]]) -> Tuple[Optional[int], Dict[str, int]]:
    best_index: Optional[int] = None
    best_mapping: Dict[str, int] = {}
    for index, row in enumerate(rows[:HEADER_SEARCH_ROWS], start=1):
        mapping = _map_headers(row)
        if len(mapping) > len(best_mapping):
            best_index = index
            best_mapping = mapping
        if all(field in mapping for field in REQUIRED_FIELDS):
            return index, mapping
    return best_index, best_mapping


def parse_external_results_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception as exc:
        return {"status": False, "error": f"Не удалось прочитать Excel: {exc}"}

    try:
        if not workbook.worksheets:
            return {"status": False, "error": "В файле нет листов"}

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {"status": False, "error": "Файл пустой"}

        header_row_index, column_map = _find_header(rows)
        missing = [field for field in REQUIRED_FIELDS if field not in column_map]
        if missing:
            return {
                "status": False,
                "error": "Не найдены обязательные колонки: ФИО и Процент правильных ответов (%).",
            }

        parsed_rows: List[Dict[str, Any]] = []
        for row_number, row in enumerate(rows[header_row_index:], start=(header_row_index or 1) + 1):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            percent_value = _parse_number(_cell_value(row, column_map.get("percent")))
            correct_value = _parse_number(_cell_value(row, column_map.get("correct_count")))
            completed_raw = _cell_value(row, column_map.get("completed_at"))
            completed_at = (
                completed_raw.isoformat(sep=" ")
                if isinstance(completed_raw, datetime)
                else _string_value(row, column_map.get("completed_at"))
            )

            parsed_rows.append(
                {
                    "row": row_number,
                    "source_number": _string_value(row, column_map.get("source_number")),
                    "platform_user": _string_value(row, column_map.get("platform_user")),
                    "ip": _string_value(row, column_map.get("ip")),
                    "completed_at": completed_at,
                    "time_spent": _string_value(row, column_map.get("time_spent")),
                    "login": _string_value(row, column_map.get("login")),
                    "full_name": _string_value(row, column_map.get("full_name")),
                    "correct_count": correct_value,
                    "percent": percent_value,
                }
            )

        if not parsed_rows:
            return {"status": False, "error": "В файле нет строк с данными"}

        return {
            "status": True,
            "sheet_name": sheet.title,
            "header_row": header_row_index,
            "rows": parsed_rows,
        }
    finally:
        workbook.close()

