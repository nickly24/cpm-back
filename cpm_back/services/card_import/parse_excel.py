"""Парсинг Excel-файла импорта карточек."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

QUESTION_HEADERS = ("вопрос",)
ANSWER_HEADERS = ("ответ",)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def _map_headers(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    normalized = [_normalize_header(cell) for cell in header_row]

    for index, header in enumerate(normalized):
        if header in QUESTION_HEADERS and "question" not in mapping:
            mapping["question"] = index
        if header in ANSWER_HEADERS and "answer" not in mapping:
            mapping["answer"] = index

    return mapping


def _cell_value(row: tuple, index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _row_is_empty(row: tuple) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def parse_cards_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return {"status": False, "error": f"Не удалось прочитать Excel: {exc}"}

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return {"status": False, "error": "Файл пустой"}

    column_map = _map_headers(list(header_row))
    if "question" not in column_map or "answer" not in column_map:
        workbook.close()
        return {
            "status": False,
            "error": "Не найдены обязательные колонки «Вопрос» и «Ответ» в первой строке.",
        }

    rows: List[Dict[str, Any]] = []
    excel_row_no = 1
    for raw_row in rows_iter:
        excel_row_no += 1
        if _row_is_empty(raw_row):
            workbook.close()
            return {
                "status": False,
                "error": f"Пустая строка {excel_row_no} недопустима. Удалите её из файла.",
            }

        question = _cell_value(raw_row, column_map.get("question"))
        answer = _cell_value(raw_row, column_map.get("answer"))
        rows.append(
            {
                "row": excel_row_no,
                "question": question,
                "answer": answer,
            }
        )

    workbook.close()

    if not rows:
        return {"status": False, "error": "В файле нет строк с карточками"}

    return {
        "status": True,
        "rows": rows,
        "sheet_name": sheet.title,
    }
